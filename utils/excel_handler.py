import openpyxl
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
import logging
from config import Config

logger = logging.getLogger(__name__)

def ensure_excel_file(filepath, headers):
    """
    Ensure Excel file exists with proper headers
    """
    if not os.path.exists(filepath):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        # Add headers
        for col, header in enumerate(headers, 1):
            sheet[f"{get_column_letter(col)}1"] = header
        
        # Save the workbook
        workbook.save(filepath)
        logger.info(f"Created new Excel file: {filepath}")

def save_user(user_data):
    """
    Save user data to Excel file
    """
    headers = ['name', 'email', 'contact', 'brief', 'created_at']
    ensure_excel_file(Config.USERS_EXCEL, headers)
    
    try:
        workbook = openpyxl.load_workbook(Config.USERS_EXCEL)
        sheet = workbook.active
        
        # Find the next empty row
        next_row = sheet.max_row + 1
        
        # Add user data
        for col, header in enumerate(headers, 1):
            sheet[f"{get_column_letter(col)}{next_row}"] = user_data.get(header, '')
        
        workbook.save(Config.USERS_EXCEL)
        logger.info(f"Saved user data for: {user_data['email']}")
        
    except Exception as e:
        logger.error(f"Error saving user data: {str(e)}")
        raise

def get_user(email):
    """
    Get user data by email
    """
    if not os.path.exists(Config.USERS_EXCEL):
        return None
    
    try:
        workbook = openpyxl.load_workbook(Config.USERS_EXCEL)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row=2):
            row_data = {headers[i]: cell.value for i, cell in enumerate(row)}
            if row_data['email'] == email:
                return row_data
        
        return None
        
    except Exception as e:
        logger.error(f"Error getting user data: {str(e)}")
        raise

def save_appointment(appointment_data):
    """
    Save appointment data to Excel file
    """
    headers = ['user_email', 'date', 'time', 'call_type', 'meeting_type', 
              'meeting_link', 'notes', 'created_at']
    ensure_excel_file(Config.APPOINTMENTS_EXCEL, headers)
    
    try:
        workbook = openpyxl.load_workbook(Config.APPOINTMENTS_EXCEL)
        sheet = workbook.active
        
        # Find the next empty row
        next_row = sheet.max_row + 1
        
        # Add appointment data
        for col, header in enumerate(headers, 1):
            sheet[f"{get_column_letter(col)}{next_row}"] = appointment_data.get(header, '')
        
        workbook.save(Config.APPOINTMENTS_EXCEL)
        logger.info(f"Saved appointment for user: {appointment_data['user_email']}")
        
    except Exception as e:
        logger.error(f"Error saving appointment data: {str(e)}")
        raise

def get_user_appointments(email):
    """
    Get all appointments for a user
    """
    if not os.path.exists(Config.APPOINTMENTS_EXCEL):
        return []
    
    try:
        workbook = openpyxl.load_workbook(Config.APPOINTMENTS_EXCEL)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        
        appointments = []
        for row in sheet.iter_rows(min_row=2):
            row_data = {headers[i]: cell.value for i, cell in enumerate(row)}
            if row_data['user_email'] == email:
                appointments.append(row_data)
        
        return appointments
        
    except Exception as e:
        logger.error(f"Error getting user appointments: {str(e)}")
        raise

def get_all_appointments():
    """
    Get all appointments
    """
    if not os.path.exists(Config.APPOINTMENTS_EXCEL):
        return []
    
    try:
        workbook = openpyxl.load_workbook(Config.APPOINTMENTS_EXCEL)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        
        appointments = []
        for row in sheet.iter_rows(min_row=2):
            appointments.append({headers[i]: cell.value for i, cell in enumerate(row)})
        
        return appointments
        
    except Exception as e:
        logger.error(f"Error getting all appointments: {str(e)}")
        raise
